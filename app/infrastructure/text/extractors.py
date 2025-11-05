"""Extracción de texto desde varios formatos comunes.

Soportados: txt, md, csv, pdf, docx, xlsx.
No se hace OCR de imágenes en esta fase.
"""
from __future__ import annotations

from io import BytesIO, StringIO
from typing import Tuple
import csv

import chardet


def _decode_bytes(data: bytes) -> str:
    if not data:
        return ""
    det = chardet.detect(data) or {}
    enc = det.get("encoding") or "utf-8"
    try:
        return data.decode(enc, errors="ignore")
    except Exception:
        return data.decode("utf-8", errors="ignore")


def extract_text_from_txt(data: bytes) -> str:
    return _decode_bytes(data)


def extract_text_from_md(data: bytes) -> str:
    # Por simplicidad: tratamos Markdown como texto plano
    return _decode_bytes(data)


def extract_text_from_csv(data: bytes) -> str:
    text = _decode_bytes(data)
    sio = StringIO(text)
    reader = csv.reader(sio)
    rows = ["\t".join(c.strip() for c in row) for row in reader]
    return "\n".join(rows)


def extract_text_from_pdf(data: bytes) -> Tuple[str, int]:
    """Extrae texto de PDF. Retorna (texto, paginas)."""
    from pypdf import PdfReader

    pdf = PdfReader(BytesIO(data))
    parts = []
    for page in pdf.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            parts.append("")
    text = "\n\n".join(p.strip() for p in parts if p is not None)
    return text, len(pdf.pages)


def extract_text_from_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(BytesIO(data))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text]
    return "\n".join(paragraphs)


def extract_text_from_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append(f"# Hoja: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            parts.append("\t".join(cells))
    return "\n".join(parts)

